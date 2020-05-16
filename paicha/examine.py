import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# 1.风险表(最好改为英文)
riskPath = "C:/Users/79234/Desktop/examine/risk.xlsx"
# 2.排查表(最好改为英文)
pcPath = "C:/Users/79234/Desktop/examine/data.xlsx"
# 3.最终生成的目录(最好改为英文)
finalPath = "C:/Users/79234/Desktop/examine/wb.xlsx"
# 获取文件

wb = load_workbook(riskPath)

wb2 = load_workbook(pcPath)

final = load_workbook(finalPath)
# 选择第一张表
riskWS = wb[wb.sheetnames[0]]
exWS = wb2[wb2.sheetnames[0]]
finalWS = final[final.sheetnames[0]]

# 拆分单元格并赋值函数
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 设置排查表格式，便于获取


def unMerged(sheet):
    m_list = sheet.merged_cells
    # 行、列
    mergedArr = []
    for m_area in m_list:
        # 合并单元格的起始坐标、终止坐标 r1-r2横向 c1-c2纵向
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 合并单元格位置信息
        if r2 - r1 > 0 or c2-c1 > 0:
            mergedArr.append((r1, r2, c1, c2))
    # 合并单元格的信息提取出再拆分
    for r in mergedArr:
        value = sheet.cell(r[0], r[2]).value
        # 拆分单元格
        sheet.unmerge_cells(
            start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])
        # 循环赋值
        for col in range(r[2], r[3]):
            # 向右赋值
            sheet.cell(row=r[0], column=col+1,
                       value=value)
        # 纵向
        for row in range(r[0], r[1]):
            # row为左闭右开区间  row+1 = (r[0]+1)~r[1]
            # 向下赋值
            sheet.cell(
                row=row+1, column=r[2], value=value)
            for col in range(r[2], r[3]):
                # 向右下赋值
                sheet.cell(row=row+1, column=col+1,
                           value=value)


# 只可以先拆分再删除
unMerged(exWS)

# 获取排查单位，提前储存值
deptName = exWS["D2"].value

exMaxCol = exWS.max_row
# 先从后面删除，不影响索引
exWS.delete_rows(exMaxCol-4, 5)
# 删除表头 从第一列起的4列的
exWS.delete_rows(1, 3)
# 删除风险等级
exWS.delete_cols(6, 1)
# 删除序号
exWS.delete_cols(1, 1)
# 措施内容
colE = column_index_from_string('E')
colQ = column_index_from_string('Q')
colS = column_index_from_string('S')


def appenRow(n, m):  # 向后追加行
    # n为当前所需复制行 dataArr为数组
    # 获取最大列数并存储当前行内值
    maxCol = exWS.max_column
    arr = []
    for c in range(1, maxCol+1):
        arr.append(exWS.cell(row=n, column=c).value)

        # 根据数据的条数 向后追加行
    for index, item in enumerate(m):
        # 0是空字符串 1是第一条字符串 追加的要是从2起
        if index >= 1:
            arr[colE-1] = item
            exWS.append(arr)


max = exWS.max_row
# 将无序列表拆分
for n in range(1, max+1):
    # 匹配存在序列
    if re.match("□", exWS.cell(row=n, column=colE).value):
        baseStr = re.sub("", "", exWS.cell(row=n, column=colE).value)
        # 换行符替换为空
        str1 = re.sub("\n", "", baseStr)
        # 将数据根据拆分为数组
        str2 = re.split("□", str1)
        # 排除空的项
        dataArr = [i for i in str2 if i != '']
        # n为当前行 dataArr为数组
        appenRow(n, dataArr)
        # # 将第一条数据赋值
        exWS.cell(row=n, column=colE, value=dataArr[0])

# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 在风险表里匹配排查的内容
# 排查表匹配赋值后删除，永远匹配第一行、最后一行
for e in exWS.values:
    # 获取风险表最后一行的内容
    for r in riskWS.values:
        if e[0] == r[0] and e[1] == r[2] and e[2] == r[3] and e[3] == r[4]:
            row = list(r)
            row[colQ-1] = "是"
            row.append("")
            row.append(deptName)
# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 在最终表里追加匹配后追加内容
            finalWS.append(row)
            break

        
final.save(finalPath)
