import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

path = "C:/Users/79234/Desktop/py/"

# 获取文件
# 1.危险辨识表(最好改为英文)
wb = load_workbook(path+'1/demo.xlsx')
# 2.风险管控表(最好改为英文)
wb2 = load_workbook(path+'2/demo.xlsx')

# 选择第一张表
worksheet = wb[wb.sheetnames[0]]
worksheet2 = wb2[wb2.sheetnames[0]]

# 拆分单元格并赋值
# 合并单元格的位置信息,可迭代对象(单个是一个'openpyxl.worksheet.cell_range.CellRange'对象)


def unMerged(sheet):
    m_list = sheet.merged_cells
    # 行、列
    mergedArr = []
    for m_area in m_list:
        # 合并单元格的起始坐标、终止坐标 r1-r2横向 c1-c2纵向
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 合并单元格位置信息
        if r2 - r1 > 0 or c2-c1 > 0:
            mergedArr.append((r1, r2, c1, c2))
    # 合并单元格的信息提取出再拆分
    for r in mergedArr:
        value = sheet.cell(r[0], r[2]).value
        # 拆分单元格
        sheet.unmerge_cells(
            start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])
        # 循环赋值
        for col in range(r[2], r[3]):
            # 向右赋值
            sheet.cell(row=r[0], column=col+1,
                       value=value)
        # 纵向
        for row in range(r[0], r[1]):
            # row为左闭右开区间  row+1 = (r[0]+1)~r[1]
            # 向下赋值
            sheet.cell(
                row=row+1, column=r[2], value=value)
            for col in range(r[2], r[3]):
                # 向右下赋值
                sheet.cell(row=row+1, column=col+1,
                           value=value)


# 只可以先拆分再删除
unMerged(worksheet)
unMerged(worksheet2)

# 删除表头 从第一列起的4列的
worksheet.delete_rows(1, 4)
worksheet2.delete_rows(1, 4)

# 创建一个工作薄对象,也就是创建一个excel文档
final = Workbook()

# 指定当前显示（活动）的sheet对象
ws = final.active
# 给新表赋值
max_row = worksheet.max_row

# 设置函数

# 复制列函数（逐行写入对性能不太好）


def setData(m, n, row, key):  # m为复制后的行数,n为所复制内容的行数,row为措施内容列数,key为措施名称
    # 赋值主要内容
    ws.cell(row=m, column=column_index_from_string('A'), value=worksheet.cell(
        row=n, column=column_index_from_string('B')).value)
    ws.cell(row=m, column=column_index_from_string('C'), value=worksheet.cell(
        row=n, column=column_index_from_string('C')).value)
    ws.cell(row=m, column=column_index_from_string('D'), value=worksheet.cell(
        row=n, column=column_index_from_string('D')).value)
    ws.cell(row=m, column=column_index_from_string('E'), value=worksheet.cell(
        row=n, column=column_index_from_string('E')).value)
    ws.cell(row=m, column=column_index_from_string('F'), value=worksheet.cell(
        row=n, column=column_index_from_string('F')).value)
    ws.cell(row=m, column=column_index_from_string('G'), value=worksheet.cell(
        row=n, column=column_index_from_string('G')).value)
    ws.cell(row=m, column=column_index_from_string('H'), value=worksheet.cell(
        row=n, column=column_index_from_string('H')).value)
    ws.cell(row=m, column=column_index_from_string('I'), value=worksheet.cell(
        row=n, column=column_index_from_string('I')).value)
    ws.cell(row=m, column=column_index_from_string('J'), value=worksheet.cell(
        row=n, column=column_index_from_string('J')).value)
    ws.cell(row=m, column=column_index_from_string('K'), value=worksheet.cell(
        row=n, column=column_index_from_string('K')).value)
    ws.cell(row=m, column=column_index_from_string('L'), value=key)
    ws.cell(row=m, column=column_index_from_string('M'), value=worksheet.cell(
        row=n, column=column_index_from_string(row)).value)
    ws.cell(row=m, column=column_index_from_string('Q'), value='否')
    # 赋值责任单位
    ws.cell(row=m, column=column_index_from_string('N'), value=worksheet2.cell(
        row=n, column=column_index_from_string('K')).value)
    ws.cell(row=m, column=column_index_from_string('O'), value=worksheet2.cell(
        row=n, column=column_index_from_string('K')).value)
    ws.cell(row=m, column=column_index_from_string('P'), value=worksheet2.cell(
        row=n, column=column_index_from_string('K')).value)


# 根据setData复制代码 设置不同的管控措施
for m in range(1, max_row+1):
    setData(m, m, 'Q', '管理措施')
    setData(m+max_row, m, 'R', '培训教育措施')
    setData(m+max_row+max_row, m, 'S', '个体防护措施')
    setData(m+max_row+max_row+max_row, m, 'T', '应急处置措施')
    setData(m+max_row+max_row+max_row, m, 'P', '工程控制措施')

wb.close()
wb2.close()

# 递归函数匹配存在表格
max = ws.max_row
# print(re.sub("\d{1,2}、", "@", ws["M1"].value).split('@'))
cols = column_index_from_string('M')


def appenRow(n, m):  # 向后追加行
    # 获取最大列数并存储当前行内值
    maxCol = ws.max_column
    arr = []
    for c in range(1, maxCol+1):
        arr.append(ws.cell(row=1, column=c).value)

        # 根据数据的条数 向后追加行
    for index, item in enumerate(m):
        # 0是空字符串 1是第一条字符串 追加的要是从2起
        if index >= 1:
            # arr[12]为措施内容列
            arr[12] = item
            ws.append(arr)

# ^1、| \n\d{1,2}、
# 改为倒叙，删除行不会对索引产生影响
for n in range(max, 0, -1):
    # 匹配存在序列
    if re.match("^1、|[2-9]{1,2}、", ws.cell(row=n, column=cols).value):
        # 换行符替换为空
        str1 = re.sub("\n", "", ws.cell(row=n, column=cols).value)
        # 将所有 1、/2、...换为@
        str2 = re.split("^1、|[2-9]{1,2}、", str1)
        dataArr = [i for i in str2 if i != '']
        # # 将数据根据@拆分为数组
        appenRow(n, dataArr)
        # # 将第一条数据赋值
        ws.cell(row=n, column=cols, value=dataArr[0])
    # 删除M列内容为"/"的行及空行
    if ws.cell(row=n, column=cols).value == "" or ws.cell(row=n, column=cols).value == "/":
        ws.delete_rows(n, 1)

final.save(path+'final.xlsx')
